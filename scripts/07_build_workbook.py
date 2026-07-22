# -*- coding: utf-8 -*-
# Builds output/MSNA_2026_sampling_frame_workbook.xlsx: a combined, styled
# Excel workbook (README + Sampling Frame + Strata-Level Summary tabs) from
# output/stage2_sampling_frame.csv and output/strata_level_sampling_frame.csv.
# Requires: pip install openpyxl
import csv
import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Derived from this script's own location (not hardcoded) so the project can
# be moved/renamed without breaking this path. This script lives in
# scripts/, one level below the project root, so output/ is a parent-level
# sibling of scripts/.
PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = PROJECT_DIR + r"\output\MSNA_2026_sampling_frame_workbook.xlsx"

SF_CSV = PROJECT_DIR + r"\output\stage2_sampling_frame.csv"
ST_CSV = PROJECT_DIR + r"\output\strata_level_sampling_frame.csv"

NAVY = "1B2A4A"
BLUE = "2C5F8A"
BLUE_LIGHT = "DCE6F1"
GREEN = "1F7A5C"
GREEN_LIGHT = "D5F0E3"
WHITE = "FFFFFF"

AMBER = "FFF2CC"
ORANGE = "FCE4D6"
PURPLE = "E6D9F2"
MINT = "D5F0E3"

# ---------------------------------------------------------------------------
# Column type coercion + widths
# ---------------------------------------------------------------------------

SF_NUMERIC = {
    "interview_number", "replacement_rank", "confidence", "building_area_m2",
    "latitude", "longitude", "x_utm", "y_utm", "households_in_cluster",
    "target_households", "selection_count", "psu_probability",
    "ssu_probability", "base_weight", "site_radius_m", "n_other_sites_in_hex",
}
SF_BOOL = {
    "certainty_stratum", "below_target_cluster", "reallocated",
    "supplementary_cluster",
}

ST_NUMERIC = {
    "n_pop", "N_hh", "n_hex", "clusters_target_stage1", "achieved_clusters",
    "m_used", "ICC", "DEFF", "expected_households_stage1", "target_sample",
    "achieved_sample", "confidence_level_pct", "target_moe_pct",
    "realized_moe_pct",
}
ST_BOOL = {"certainty_stratum"}

SF_WIDTHS = {
    "survey_id": 24, "cluster_id": 20, "status": 10, "interview_number": 10,
    "replacement_rank": 10, "pop_type": 10, "strata_id": 18, "region": 8,
    "adm1_pcode": 11, "adm1_name": 14, "adm2_pcode": 11, "adm2_name": 16,
    "adm3_pcode": 12, "adm3_name": 18, "admin3_source": 14,
    "admin3_cod_pcode": 14, "admin3_cod_name": 18, "uuid_hex": 22,
    "uuid": 30, "building_id": 16, "confidence": 11, "building_area_m2": 14,
    "latitude": 11, "longitude": 11, "x_utm": 12, "y_utm": 12,
    "households_in_cluster": 12, "target_households": 12,
    "selection_count": 11, "certainty_stratum": 12, "selection_type": 12,
    "psu_probability": 12, "ssu_probability": 12, "base_weight": 11,
    "below_target_cluster": 13, "reallocated": 11,
    "original_uuid_hex_pop": 20, "supplementary_cluster": 14,
    "location_source": 16, "households_in_cluster_source": 20,
    "site_radius_m": 11, "iom_site_id": 22, "iom_site_name": 22,
    "iom_site_type": 14, "iom_site_ward": 16, "n_other_sites_in_hex": 12,
    "idp_population_category": 18,
}
ST_WIDTHS = {
    "region": 8, "adm1_pcode": 11, "adm1_name": 14, "adm2_pcode": 11,
    "adm2_name": 16, "pop_type": 8, "strata_id": 18, "n_pop": 11, "N_hh": 10, "n_hex": 8,
    "selection_type": 12, "certainty_stratum": 12,
    "clusters_target_stage1": 14, "achieved_clusters": 13, "m_used": 8,
    "ICC": 7, "DEFF": 7, "expected_households_stage1": 16,
    "target_sample": 12, "achieved_sample": 13, "confidence_level_pct": 13,
    "target_moe_pct": 12, "realized_moe_pct": 13,
}


def coerce(value, header, numeric_set, bool_set):
    # R's write_csv() writes missing values as the literal text "NA" - must
    # be treated as blank here too, or numeric/boolean columns end up as
    # mixed-type text in Excel (breaks sort/filter/sum on those columns).
    if value is None or value == "" or value == "NA":
        return None
    if header in bool_set:
        if value == "TRUE":
            return True
        if value == "FALSE":
            return False
        return None
    if header in numeric_set:
        try:
            f = float(value)
            if f.is_integer():
                return int(f)
            return f
        except ValueError:
            return value
    return value


def write_data_sheet(wb, sheet_name, csv_path, numeric_set, bool_set,
                      widths, header_color, table_style, note):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{sheet_name} — see README tab for column definitions and legend"
    ws["A1"].font = Font(bold=True, italic=True, size=10, color="595959")
    ws["A2"] = note
    ws["A2"].font = Font(italic=True, size=9, color="7F7F7F")

    start_row = 4
    with open(csv_path, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        for i, h in enumerate(header, start=1):
            ws.cell(row=start_row, column=i, value=h)
        n_rows = 0
        for row in reader:
            coerced = [
                coerce(v, h, numeric_set, bool_set)
                for v, h in zip(row, header)
            ]
            ws.append(coerced)
            n_rows += 1

    last_col_letter = get_column_letter(len(header))
    last_row = start_row + n_rows

    for i, h in enumerate(header, start=1):
        w = widths.get(h, 13)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{start_row + 1}"

    tbl_ref = f"A{start_row}:{last_col_letter}{last_row}"
    table = Table(displayName=sheet_name.replace(" ", "_").replace("-", "_"), ref=tbl_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    for i in range(1, len(header) + 1):
        ws.cell(row=start_row, column=i).font = Font(bold=True, color=WHITE)
        ws.cell(row=start_row, column=i).fill = PatternFill(
            "solid", fgColor=header_color
        )

    return ws, header, start_row, last_row, last_col_letter


def add_flag_highlight(ws, header, start_row, last_row, last_col_letter, flag_col, fill_hex):
    col_idx = header.index(flag_col) + 1
    col_letter = get_column_letter(col_idx)
    rng = f"A{start_row + 1}:{last_col_letter}{last_row}"
    rule = FormulaRule(
        formula=[f"${col_letter}{start_row + 1}=TRUE"],
        fill=PatternFill("solid", fgColor=fill_hex),
    )
    ws.conditional_formatting.add(rng, rule)


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- Sampling Frame sheet -----------------------------------------------
ws_sf, sf_header, sf_start, sf_last, sf_last_col = write_data_sheet(
    wb, "Sampling Frame", SF_CSV, SF_NUMERIC, SF_BOOL, SF_WIDTHS,
    BLUE, "TableStyleMedium2",
    "One row per planned interview (primary or reserve), Non-IDP and IDP combined.",
)
add_flag_highlight(ws_sf, sf_header, sf_start, sf_last, sf_last_col, "below_target_cluster", AMBER)
add_flag_highlight(ws_sf, sf_header, sf_start, sf_last, sf_last_col, "reallocated", ORANGE)
add_flag_highlight(ws_sf, sf_header, sf_start, sf_last, sf_last_col, "supplementary_cluster", PURPLE)

print(f"Sampling Frame: {sf_last - sf_start} data rows written")

# --- Strata-Level Summary sheet ------------------------------------------
ws_st, st_header, st_start, st_last, st_last_col = write_data_sheet(
    wb, "Strata-Level Summary", ST_CSV, ST_NUMERIC, ST_BOOL, ST_WIDTHS,
    GREEN, "TableStyleMedium7",
    "One row per population-group x LGA stratum - for validation and reporting.",
)
add_flag_highlight(ws_st, st_header, st_start, st_last, st_last_col, "certainty_stratum", MINT)

print(f"Strata-Level Summary: {st_last - st_start} data rows written")

# ---------------------------------------------------------------------------
# README sheet
# ---------------------------------------------------------------------------

SF_DEFS = [
    ("Identifiers", None),
    ("survey_id", "Unique identifier for this specific interview slot (primary or reserve)."),
    ("cluster_id", "Identifier for the cluster (hexagon or IDP site) this interview belongs to."),
    ("status", "Whether this row is a “primary” interview to conduct first, or a “reserve” replacement if a primary household can’t be reached."),
    ("interview_number", "The primary household’s sequence number within its cluster (1 to target); blank for reserve rows."),
    ("replacement_rank", "The reserve household’s order of use within its cluster if a primary needs replacing; blank for primary rows."),
    ("pop_type", "Population group this record belongs to: “non_idp” (not internally displaced) or “idp” (internally displaced)."),
    ("strata_id", "Identifier for the sampling stratum (population group + LGA) this cluster was drawn from."),
    ("Geography", None),
    ("region", "Nigeria’s North-East / North-West / North-Central region grouping for this location."),
    ("adm1_pcode", "State-level administrative P-code."),
    ("adm1_name", "State name."),
    ("adm2_pcode", "LGA (Local Government Area)-level administrative P-code — the stratification unit."),
    ("adm2_name", "LGA name."),
    ("adm3_pcode", "Ward-level administrative P-code, from the GRID3 ward boundary layer."),
    ("adm3_name", "Ward name, from the GRID3 ward boundary layer."),
    ("admin3_source", "Which boundary layer (GRID3 ward or COD Admin-3) was used to assign this record’s ward/admin-3 attribution."),
    ("admin3_cod_pcode", "Ward/Admin-3 P-code from the official COD Admin-3 layer (only available in the three North-East states), as a secondary reference."),
    ("admin3_cod_name", "Ward/Admin-3 name from the official COD Admin-3 layer, as a secondary reference."),
    ("Location", None),
    ("uuid_hex", "Identifier for the selected hexagon (the Stage 1 primary sampling unit) this record belongs to."),
    ("uuid", "Identifier for the specific building footprint (Non-IDP) or IOM DTM site (IDP) this interview location was drawn from."),
    ("building_id", "Google Open Buildings identifier for the selected building footprint (Non-IDP records only)."),
    ("confidence", "Google Open Buildings confidence score for the selected building footprint (Non-IDP records only)."),
    ("building_area_m2", "Footprint area in square metres of the selected building (Non-IDP records only)."),
    ("latitude", "Latitude (WGS84) of the interview location."),
    ("longitude", "Longitude (WGS84) of the interview location."),
    ("x_utm", "Projected easting coordinate of the interview location, in metres."),
    ("y_utm", "Projected northing coordinate of the interview location, in metres."),
    ("location_source", "Whether the interview location came from a building footprint (“building_footprint”) or an IOM DTM site GPS point (“idp_site_point”)."),
    ("Design and weighting", None),
    ("households_in_cluster", "Total number of eligible households identified in this cluster (deduplicated building count for Non-IDP; DTM-reported estimate for IDP)."),
    ("households_in_cluster_source", "Whether households_in_cluster is a deduplicated building count or a provisional DTM estimate."),
    ("target_households", "Number of households this cluster was intended to deliver (accounts for repeated PPS draws)."),
    ("selection_count", "Number of times this cluster’s hexagon was drawn in the Stage 1 systematic PPS selection."),
    ("certainty_stratum", "Whether this record’s stratum was fully enumerated (every eligible hexagon included) rather than sampled by probability."),
    ("selection_type", "How this cluster was selected: “certainty” (full enumeration) or “pps” (probability-proportional-to-size)."),
    ("psu_probability", "The cluster’s (Stage 1) probability of selection."),
    ("ssu_probability", "The household’s (Stage 2) probability of selection within its cluster."),
    ("base_weight", "The design weight for this record: 1 ÷ (psu_probability × ssu_probability)."),
    ("Flags", None),
    ("below_target_cluster", "Flags a cluster where fewer eligible households were available than the target, so it delivers everything available rather than the full target."),
    ("reallocated", "Flags a Non-IDP cluster that was swapped to a different hexagon in the same stratum because the original had no eligible buildings."),
    ("original_uuid_hex_pop", "The originally-selected hexagon’s identifier, kept for reference when a cluster has been reallocated."),
    ("supplementary_cluster", "Flags a cluster that was added after the original Stage 1 draw to close a stratum’s shortfall against its target_sample."),
    ("IDP site detail", None),
    ("site_radius_m", "Fixed radius in metres field teams should treat as “this site” around an IDP site’s GPS point (IDP records only)."),
    ("iom_site_id", "IOM DTM site identifier for the selected IDP site (IDP records only)."),
    ("iom_site_name", "IOM DTM site name for the selected IDP site (IDP records only)."),
    ("iom_site_type", "IOM DTM site type classification for the selected IDP site (IDP records only)."),
    ("iom_site_ward", "Ward recorded by IOM DTM for the selected IDP site (IDP records only)."),
    ("n_other_sites_in_hex", "Number of additional, smaller IOM DTM sites also located in this cluster’s hexagon beyond the one selected as representative."),
    ("idp_population_category", "Which IOM DTM Population Category group this IDP site belongs to: “idps in camp” or “idps in host” (IDP records only; returnee-classified sites are excluded from the frame entirely)."),
]

ST_DEFS = [
    ("Geography", None),
    ("region", "Nigeria’s North-East / North-West / North-Central region grouping for this stratum."),
    ("adm1_pcode", "State-level administrative P-code."),
    ("adm1_name", "State name."),
    ("adm2_pcode", "LGA-level administrative P-code — the stratification unit."),
    ("adm2_name", "LGA name."),
    ("pop_type", "Population group this stratum covers: “non_idp” or “idp”."),
    ("strata_id", "Identifier for this stratum (population group + LGA) — matches strata_id in the Sampling Frame sheet for joining across tables."),
    ("Population and design", None),
    ("n_pop", "Estimated total population in this stratum."),
    ("N_hh", "Estimated total number of households in this stratum — the sampling frame’s population size."),
    ("n_hex", "Number of eligible hexagons available in this stratum before selection."),
    ("selection_type", "How this stratum was sampled: “certainty” (full enumeration) or “pps” (probability-proportional-to-size)."),
    ("certainty_stratum", "Whether this stratum was fully enumerated rather than sampled by probability."),
    ("ICC", "Intra-cluster correlation coefficient assumed for this stratum’s design effect calculation."),
    ("DEFF", "Design effect applied to this stratum’s sample size calculation, derived from m_used and ICC."),
    ("m_used", "Households-per-cluster target actually used for this stratum (7 in the 10 boosted strata, 6 elsewhere)."),
    ("Clusters and sample size", None),
    ("clusters_target_stage1", "Number of clusters the Stage 1 sample size formula calculated for this stratum."),
    ("achieved_clusters", "Number of clusters actually delivered in this stratum, including any reallocated or supplementary clusters."),
    ("expected_households_stage1", "Household sample size originally calculated by the Stage 1 formula, before any boosting or reallocation."),
    ("target_sample", "The stratum’s true design target: Stage-1 cluster count × m_used (or full population for certainty strata)."),
    ("achieved_sample", "Number of primary household interviews actually delivered in this stratum."),
    ("Precision", None),
    ("confidence_level_pct", "Confidence level assumed for this stratum’s margin-of-error calculation (90%)."),
    ("target_moe_pct", "Target margin of error for this stratum (10% for PPS strata; not applicable for certainty strata)."),
    ("realized_moe_pct", "Margin of error actually achieved, calculated from the delivered sample size rather than the target."),
]


def write_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100

    row = 1

    def title(text, size=16, color=NAVY):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=size, color=color)
        row += 2

    def para(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        c.font = Font(size=11)
        row += 2

    def subhead(text, color):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c.font = Font(bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=color)
        row += 1

    def group_row(text, color):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c.font = Font(bold=True, italic=True, size=10, color=color)
        row += 1

    def def_row(col_name, definition):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=col_name)
        c1.font = Font(name="Consolas", size=10, bold=True)
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2 = ws.cell(row=row, column=2, value=definition)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    title("MSNA 2026 — Sampling Frame Workbook")

    para(
        "This workbook accompanies the sampling frame for the 2026 Multi-Sector Needs "
        "Assessment (MSNA) covering North-West, North-East, and North-Central Nigeria. "
        "The design is a stratified two-stage cluster sample, stratified by population "
        "group (Non-IDP / IDP) and LGA. Non-IDP clusters are drawn by probability-"
        "proportional-to-size (PPS) selection weighted by gridded population, with Stage "
        "2 household locations drawn from Google Open Buildings footprints. IDP clusters "
        "are drawn by PPS weighted by IOM Displacement Tracking Matrix (DTM) reported "
        "population, with interview locations anchored to each selected site’s own GPS "
        "point rather than a building draw. Non-IDP strata whose realized sample fell "
        "short of their calculated target — concentrated in but not limited to conflict-"
        "affected North-East LGAs, where building-footprint undercounting is most severe — "
        "receive the minimum number of additional supplementary clusters (drawn via the "
        "same PPS mechanism, at the standard cluster size) needed to close the gap — see "
        "the accompanying methodology document for full detail."
    )

    subhead("How to use this workbook", NAVY)
    row += 1
    def_row("Sampling Frame", "One row per planned interview (primary or reserve), Non-IDP and IDP combined — for field teams and household/interview-level analysis.")
    def_row("Strata-Level Summary", "One row per population-group × LGA stratum — for validation and reporting on sample sizes, design effects, and realized margins of error.")
    row += 1

    subhead("Highlight legend", NAVY)
    row += 1

    def legend_row(color, text):
        nonlocal row
        c = ws.cell(row=row, column=1, value="")
        c.fill = PatternFill("solid", fgColor=color)
        c2 = ws.cell(row=row, column=2, value=text)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    legend_row(AMBER, "Sampling Frame — below_target_cluster = TRUE (cluster delivered fewer households than its target).")
    legend_row(ORANGE, "Sampling Frame — reallocated = TRUE (cluster swapped to a different hexagon after the original had no eligible buildings).")
    legend_row(PURPLE, "Sampling Frame — supplementary_cluster = TRUE (cluster added after the original Stage 1 draw to close a shortfall).")
    legend_row(MINT, "Strata-Level Summary — certainty_stratum = TRUE (every eligible hexagon in this stratum was enumerated, not sampled).")
    row += 1

    subhead("Sampling Frame — column definitions", BLUE)
    row += 1
    for name, definition in SF_DEFS:
        if definition is None:
            group_row(name, BLUE)
        else:
            def_row(name, definition)
    row += 1

    subhead("Strata-Level Summary — column definitions", GREEN)
    row += 1
    for name, definition in ST_DEFS:
        if definition is None:
            group_row(name, GREEN)
        else:
            def_row(name, definition)


write_readme(wb)
wb.active = 0

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
